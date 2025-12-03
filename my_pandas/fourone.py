# -*- coding：utf-8 -*-
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles.borders import Border, Side

pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
pd.set_option('display.width', 1000)

# 将数据写入带表头的模板
# 读取带表头的模板
model_path = r"D:\社保公积金汇总表.xlsx"
wb = load_workbook(model_path)
ws = wb.active
count = 0
folder_path = "Z:\附件(1)"
file_path = []
# 特殊地区列表
SPECIAL_REGIONS = ["陕西", "惠州", "深圳", "西安", "韶关"]
df_total = pd.DataFrame()
for root, dirs, files in os.walk(folder_path):
    for file in files:
        file_path = os.path.join(root, file)
        area = file_path.split("\\")[-1].split("-")[-1]
        if area in ["公积金缴费明细.xlsx", "社保缴费明细.xlsx"]:
            print(file_path)
            continue
        # pandas读取单个excel文件
        # file_path = r"Z:\附件(1)\东莞-东莞市必达能源供应链有限公司-社保公积金缴费明细.xlsx"
        data_df = pd.read_excel(file_path, sheet_name=0, skiprows=4, header=None)
        # 原文件数据设置字段名
        area = file_path.split("\\")[-1].split("-")[0]
        if area not in SPECIAL_REGIONS :
            data_df.columns = ["序号", "姓名", "社保_总合计", "社保_单位合计", "社保_个人合计", "养老保险_单位交",
                               "养老保险_个人交", "工伤保险_单位交", "工伤保险_个人交", "失业保险_单位交",
                               "失业保险_个人交",
                               "基础医疗_单位交", "基础医疗_个人交", "公积金_单位交", "公积金_个人交", "公积金_合计",
                               "部门",
                               "备注"]
            # 增加文件字段名并设置空值
            data_df["补充医疗_单位交"] = None
            data_df["补充医疗_个人交"] = None
        else:
            data_df.columns = ["序号", "姓名", "社保_总合计", "社保_单位合计", "社保_个人合计", "养老保险_单位交",
                               "养老保险_个人交", "工伤保险_单位交", "工伤保险_个人交", "失业保险_单位交",
                               "失业保险_个人交", "基础医疗_单位交", "基础医疗_个人交", "补充医疗_单位交", "补充医疗_个人交",
                               "公积金_单位交", "公积金_个人交", "公积金_合计", "部门", "备注"]

        # 调整字段顺序
        data_df_new = data_df[
            ["姓名", "社保_总合计", "社保_单位合计", "社保_个人合计", "养老保险_单位交", "养老保险_个人交",
             "工伤保险_单位交", "工伤保险_个人交", "失业保险_单位交", "失业保险_个人交", "基础医疗_单位交",
             "基础医疗_个人交",
             "补充医疗_单位交", "补充医疗_个人交", "公积金_单位交", "公积金_个人交", "公积金_合计", "部门", "备注"]]
        df_total = pd.concat([df_total, data_df_new], axis=0)

file_path1 = "Z:\附件(1)\深圳-深圳市同仁科技实业有限公司-公积金缴费明细.xlsx"
file_path2 = "Z:\附件(1)\深圳-深圳市同仁科技实业有限公司-社保缴费明细.xlsx"

data1_df = pd.read_excel(file_path1, sheet_name=0, skiprows=4, header=None)
data2_df = pd.read_excel(file_path2, sheet_name=0, skiprows=4, header=None)

data1_df.columns = ["序号", "姓名", "社保_总合计", "社保_单位合计", "社保_个人合计", "养老保险_单位交",
                    "养老保险_个人交", "工伤保险_单位交", "工伤保险_个人交", "失业保险_单位交",
                    "失业保险_个人交",
                    "基础医疗_单位交", "基础医疗_个人交", "补充医疗_单位交", "补充医疗_个人交",
                    "公积金_单位交", "公积金_个人交", "公积金_合计", "部门", "备注"]
data2_df.columns = ["序号", "姓名", "社保_总合计", "社保_单位合计", "社保_个人合计", "养老保险_单位交",
                    "养老保险_个人交", "工伤保险_单位交", "工伤保险_个人交", "失业保险_单位交",
                    "失业保险_个人交",
                    "基础医疗_单位交", "基础医疗_个人交", "补充医疗_单位交", "补充医疗_个人交",
                    "公积金_单位交", "公积金_个人交", "公积金_合计", "部门", "备注"]

merge_list1 = data1_df[["姓名", "公积金_单位交", "公积金_个人交", "公积金_合计"]]

merge_list2 = data2_df[
    ["姓名", "社保_总合计", "社保_单位合计",
     "社保_个人合计", "养老保险_单位交", "养老保险_个人交", "工伤保险_单位交", "工伤保险_个人交",
     "失业保险_单位交",
     "失业保险_个人交", "基础医疗_单位交", "基础医疗_个人交", "补充医疗_单位交", "补充医疗_个人交",  "部门", "备注"]]

total_list = pd.merge(merge_list2, merge_list1, on=["姓名"], how="left")


# 调整字段顺序
data_df_new = total_list[
    ["姓名", "社保_总合计", "社保_单位合计", "社保_个人合计", "养老保险_单位交", "养老保险_个人交",
     "工伤保险_单位交", "工伤保险_个人交", "失业保险_单位交", "失业保险_个人交", "基础医疗_单位交",
     "基础医疗_个人交", "补充医疗_单位交", "补充医疗_个人交", "公积金_单位交", "公积金_个人交", "公积金_合计", "部门", "备注"]]
emp_list = pd.read_excel(r"D:\全部在职_20251202091012.xlsx", sheet_name=0, dtype="str")

# 横向join,使用pd.merge()函数,关联两个dataframe数据
merge_list = pd.merge(total_list, emp_list, left_on=["姓名"], right_on=["姓名"], how="left")
merge_list1 = merge_list[
    ["姓名", "社保_总合计", "社保_单位合计",
     "社保_个人合计", "养老保险_单位交", "养老保险_个人交", "工伤保险_单位交", "工伤保险_个人交",
     "失业保险_单位交", "失业保险_个人交", "基础医疗_单位交", "基础医疗_个人交", "补充医疗_单位交", "补充医疗_个人交", "公积金_单位交",
     "公积金_个人交", "公积金_合计","部门", "备注"]]
print(merge_list1.head(10))

df_total = pd.concat([df_total, merge_list1], axis=0)
df_total.insert(0, '序号', range(1, len(df_total) + 1))
print(df_total)


emp_list = pd.read_excel(r"D:\全部在职_20251202091012.xlsx", sheet_name=0, dtype="str")
theme_list = pd.read_excel(r"D:\主体维度表.xlsx", sheet_name=0, dtype="str")
empty_list = pd.read_excel(r"D:\员工部门维度表.xlsx", sheet_name=0, dtype="str")
merge_list = pd.merge(empty_list, emp_list, left_on=["实际部门", "姓名"], right_on=["组织全称", "姓名"], how="left")
merge_list = pd.merge(merge_list, theme_list, on=["公积金主体", "费用承担部门"], how="left")
df_total = pd.merge(df_total, merge_list, left_on=['姓名', '部门'], right_on=['姓名', '费用承担部门'], how="left")
df_total = df_total[
    ["序号", "费用承担公司", "费用承担部门", "组织全称", "职位", "工号", "姓名", "社保_总合计", "社保_单位合计",
     "社保_个人合计", "养老保险_单位交", "养老保险_个人交", "工伤保险_单位交", "工伤保险_个人交",
     "失业保险_单位交", "失业保险_个人交", "基础医疗_单位交", "基础医疗_个人交", "补充医疗_单位交", "补充医疗_个人交",
     "公积金_单位交", "公积金_个人交", "公积金_合计", "备注_x"]]
for value in dataframe_to_rows(df_total, index=False, header=False):
    ws.append(value)

max_cols = ws.max_column
# 定义边框样式
thin_border = Side(style="thin")
border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

# 遍历指定区域的所有单元格并应用边框
for value in range(5, 5 + len(merge_list1)):  # 第3行到第9行
    for col in range(1, max_cols + 1):  # 第1列到第19列
        cell = ws.cell(row=value, column=col)
        cell.border = border
        count += 1

print(count)
wb.save(r"D:\社保公积金缴费明细汇总.xlsx")
wb.close()
